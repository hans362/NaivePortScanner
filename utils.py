import re
from scapy.all import TCP, IP, sr1, RandShort, ICMP, ARP, UDP
from multiprocessing.pool import ThreadPool
from datetime import datetime
from openpyxl import Workbook, styles
import consts
import socket
import os
import ipaddress


def CheckPermission():
    if os.name == "nt":
        import ctypes

        if ctypes.windll.shell32.IsUserAnAdmin():
            return True
        else:
            return False
    elif os.name == "posix":
        if os.getuid() == 0:
            return True
        else:
            return False
    else:
        return False


def CidrToIps(cidr):
    ips = []
    try:
        for ip in ipaddress.ip_network(cidr):
            ips.append(str(ip))
    except ValueError:
        pass
    return ips


def ValidateTimeout(timeout):
    if len(timeout) > 5 or not timeout.isdigit():
        return False
    try:
        timeout = int(timeout)
        if timeout < 100:
            return False
        consts.SCAN_TIMEOUT = timeout
        return True
    except Exception:
        return False


def IcmpEchoScan(ip):
    response = sr1(
        IP(dst=ip) / ICMP(), timeout=consts.SCAN_TIMEOUT / 1000, verbose=False
    )
    if response:
        return ip, consts.HOST_UP
    return ip, consts.HOST_DOWN


def ArpScan(ip):
    response = sr1(ARP(pdst=ip), timeout=consts.SCAN_TIMEOUT / 1000, verbose=False)
    if response:
        return ip, consts.HOST_UP
    return ip, consts.HOST_DOWN


def ParallelHostScan(ips, callback, scan_type=IcmpEchoScan):
    pool = ThreadPool(os.cpu_count() * 4 + 1)
    for ip in ips:
        pool.apply_async(scan_type, args=(ip,), callback=callback)
    pool.close()
    return pool


def TcpConnectScan(ip, port):
    s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    s.settimeout(consts.SCAN_TIMEOUT / 1000)
    response = s.connect_ex((ip, port))
    s.close()
    if response == 0:
        s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        s.settimeout(consts.SCAN_TIMEOUT / 1000)
        s.connect_ex((ip, port))
        try:
            s.sendall(
                b"GET / HTTP/1.1\r\n"
                + f"Host: {ip}:{port}\r\n".encode("utf-8")
                + b"Connection: close\r\n"
                + b"\r\n\r\n"
            )
            banner = s.recv(1024).decode("utf-8", errors="ignore")
            title = re.search(r"<title>(.*?)</title>", banner, re.IGNORECASE)
            banner = banner.splitlines()[0].strip()
            if title:
                title = title.group(1)
            else:
                title = None
        except Exception:
            banner = None
            title = None
            pass
        s.close()
        return ip, port, consts.PORT_OPEN, banner, title
    return ip, port, consts.PORT_CLOSED | consts.PORT_FILTERED


def TcpSynScan(ip, port):
    sport = RandShort()
    response = sr1(
        IP(dst=ip) / TCP(sport=sport, dport=port, flags="S"),
        timeout=consts.SCAN_TIMEOUT / 1000,
        verbose=False,
    )
    if response:
        if response.haslayer(TCP):
            flags = response.getlayer(TCP).flags
            if "S" in flags and "A" in flags:
                return ip, port, consts.PORT_OPEN
            elif "R" in flags:
                return ip, port, consts.PORT_CLOSED
    return ip, port, consts.PORT_FILTERED


def TcpFinScan(ip, port):
    sport = RandShort()
    response = sr1(
        IP(dst=ip) / TCP(sport=sport, dport=port, flags="F"),
        timeout=consts.SCAN_TIMEOUT / 1000,
        verbose=False,
    )
    if response:
        if response.haslayer(TCP):
            flags = response.getlayer(TCP).flags
            if "R" in flags:
                return ip, port, consts.PORT_CLOSED
    return ip, port, consts.PORT_OPEN | consts.PORT_FILTERED


def UdpScan(ip, port):
    sport = RandShort()
    response = sr1(
        IP(dst=ip) / UDP(sport=sport, dport=port),
        timeout=consts.SCAN_TIMEOUT / 1000,
        verbose=False,
    )
    if response:
        if response.haslayer(UDP):
            return ip, port, consts.PORT_OPEN
        elif response.haslayer(ICMP):
            icmp_type = response.getlayer(ICMP).type
            icmp_code = response.getlayer(ICMP).code
            if icmp_type == 3 and icmp_code == 3:
                return ip, port, consts.PORT_CLOSED
            else:
                return ip, port, consts.PORT_FILTERED
    return ip, port, consts.PORT_OPEN | consts.PORT_FILTERED


def ParallelPortScan(ips, ports, callback, scan_type=TcpConnectScan):
    pool = ThreadPool(os.cpu_count() * 4 + 1)
    for ip in ips:
        for port in ports:
            pool.apply_async(scan_type, args=(ip, port), callback=callback)
    pool.close()
    return pool


def SaveResults(tree):
    filename = f"results_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["主机IP地址", "传输层协议", "端口号", "状态", "额外信息"])
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 100
    hosts = tree.get_children()
    for host in hosts:
        host_ip = tree.item(host)["text"]
        results = tree.get_children(host)
        for result in results:
            result_name = tree.item(result)["text"]
            ports = tree.get_children(result)
            for port in ports:
                port_number = tree.item(port)["text"]
                try:
                    extra_info = " ".join(
                        [tree.item(x)["text"] for x in tree.get_children(port)]
                    )
                    ws.append(
                        [
                            host_ip,
                            port_number.split("/")[1],
                            port_number.split("/")[0],
                            result_name,
                            extra_info,
                        ]
                    )
                except Exception:
                    extra_info = " ".join(
                        [
                            str(tree.item(x)["text"].encode())[2:-1]
                            for x in tree.get_children(port)
                        ]
                    )
                    ws.append(
                        [
                            host_ip,
                            port_number.split("/")[1],
                            port_number.split("/")[0],
                            result_name,
                            extra_info,
                        ]
                    )
    ws.auto_filter.ref = ws.dimensions
    for row in ws.iter_rows(
        min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column
    ):
        for cell in row:
            cell.border = cell.border + styles.Border(
                left=styles.Side(style="thin", color="000000"),
                right=styles.Side(style="thin", color="000000"),
                top=styles.Side(style="thin", color="000000"),
                bottom=styles.Side(style="thin", color="000000"),
            )
    wb.save(filename)
    return filename


if __name__ == "__main__":
    print(ParallelHostScan(CidrToIps("111.186.58.0/24")))
    # print(ParallelPortScan("111.186.58.123", [80, 443, 3389, 21, 22, 23, 25565]))
    # print(
    #     ParallelPortScan(
    #         "111.186.58.123", [80, 443, 3389, 21, 22, 23, 25565], TcpSynScan
    #     )
    # )
    # print(
    #     ParallelPortScan(
    #         "111.186.58.123", [80, 443, 3389, 21, 22, 23, 25565], TcpFinScan
    #     )
    # )
